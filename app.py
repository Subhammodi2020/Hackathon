import base64
from flask import make_response, send_file, url_for
from PIL import Image, ImageDraw, ImageFont
import qrcode
import io
from openpyxl import load_workbook
import smtplib
import qrcode
import io
import os
from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook
from datetime import datetime
from firebase_config import initialize_firebase
from firebase_config import db
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.utils import formataddr

app = Flask(__name__)
# app.config['EXCEL_FILE'] = 'employees.xlsx'  # Path to your Excel file
app.config['EXCEL_FILE'] = 'Dummy.xlsx'
app.config['MAIL_SERVER'] = 'smtp.gmail.com'  # Update with your SMTP server
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'sahil.singh@go-yubi.com'  # Update with your email
app.config['MAIL_PASSWORD'] = 'ibun alzg ldoz jfgx'  # Use App Password for Gmail
app.config['MAIL_DEFAULT_SENDER'] = 'sahil.singh@go-yubi.com'

if db is None:
    print("WARNING: Firebase not initialized. Some features may not work.")


# In-memory storage for employee data
employees = {}

# def load_employees():
#     print(f"Load employee data from Excel file")
#     try:
#         wb = load_workbook(filename=app.config['EXCEL_FILE'], read_only=True)
#         ws = wb.active
        
#         # Get headers
#         headers = [str(cell.value).lower() if cell.value else '' for cell in next(ws.rows)]
        
#         # Load data
#         for row in ws.iter_rows(min_row=2):  # Skip header
#             if not row[0].value:  # Skip empty rows
#                 continue
                
#             employee = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
#             print(f"employee: {employee}")
#             employees[employee.get('employee id')] = employee
        
#         print(f"employees: {employees}")
#         print(f"Loaded {len(employees)} employees from Excel")
#         return True
        
#     except Exception as e:
#         print(f"Error loading Excel file: {e}")
#         return False


# this was working except firebase
# def load_employees():
#     if db is None:
#         print("Firebase not initialized. Cannot load employees.")
#         return False

#     try:
#         wb = load_workbook(filename=app.config['EXCEL_FILE'], read_only=True)
#         ws = wb.active
        
#         # Get headers
#         headers = [str(cell.value).lower() if cell.value else '' for cell in next(ws.rows)]
        
#         # Clear existing data (optional)
#         # Note: Be careful with this in production
#         # docs = db.collection('employees').stream()
#         # for doc in docs:
#         #     doc.reference.delete()
        
#         # Load data
#         for row in ws.iter_rows(min_row=2):  # Skip header
#             if not row[0].value:  # Skip empty rows
#                 continue
                
#             employee_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
#             employee_id = str(employee_data.get('employee id', ''))
#             print(f"employee_id: {employee_id}")
            
#             if employee_id:
#                 # Generate QR code
#                 qr = qrcode.QRCode(version=1, box_size=10, border=5)
#                 qr.add_data(employee_id)
#                 qr.make(fit=True)
#                 img = qr.make_image(fill_color="black", back_color="white")
                
#                 # Convert to base64
#                 img_io = io.BytesIO()
#                 img.save(img_io, 'PNG')
#                 img_io.seek(0)
#                 qr_base64 = base64.b64encode(img_io.getvalue()).decode('utf-8')
#                 print(f"qr_base64: {qr_base64}")
                
#                 # Add to Firestore
#                 employee_ref = db.collection('employees').document(employee_id)
#                 print("empl")
#                 employee_ref.set({
#                     'name': employee_data.get('name', ''),
#                     'department': employee_data.get('department', ''),
#                     'position': employee_data.get('position', ''),
#                     'email': employee_data.get('email', ''),
#                     'qr_code': qr_base64,
#                     'last_updated': datetime.utcnow()
#                 })

#         print(f"Successfully loaded employees into Firestore")
#         return True
        
#     except Exception as e:
#         print(f"Error loading employees: {str(e)}")
#         return False

def send_email(recipient_email, employee_name, qr_image_data, employee_id):
    try:
        # Create message container
        msg = MIMEMultipart('related')
        msg['Subject'] = f'Your Employee QR Code - {employee_name}'
        msg['From'] = formataddr(('Your Company Name', app.config['MAIL_USERNAME']))
        msg['To'] = recipient_email

        # Create the body of the message
        html = f"""
        <html>
          <body>
            <p>Dear {employee_name},</p>
            <p>Please find your employee QR code attached. This QR code contains your employee ID: <strong>{employee_id}</strong>.</p>
            <p>You can use this QR code for attendance and other company services.</p>
            <p>Best regards,<br>Your Company</p>
          </body>
        </html>
        """
        
        # Attach HTML content
        msg.attach(MIMEText(html, 'html'))
        
        # Attach QR code image
        qr_image = MIMEImage(qr_image_data)
        qr_image.add_header('Content-ID', '<qrcode>')
        qr_image.add_header('Content-Disposition', 'attachment', filename=f'qrcode_{employee_id}.png')
        msg.attach(qr_image)
        
        # Connect to SMTP server and send email
        with smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT']) as server:
            server.starttls()
            server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
            server.send_message(msg)
            
        print(f"Email sent to {recipient_email} for employee ID: {employee_id}")
        return True
        
    except Exception as e:
        print(f"Error sending email to {recipient_email}: {str(e)}")
        return False

def load_employees():
    try:
        # Load the workbook
        wb = load_workbook(filename=app.config['EXCEL_FILE'])
        ws = wb.active
        
        # Get headers
        headers = [str(cell.value).lower() if cell.value else '' for cell in next(ws.rows)]
        print(headers)
        
        # Add QR code column if it doesn't exist
        if 'qr code' not in headers:
            headers.append('QR Code')
            ws.cell(row=1, column=len(headers), value='QR Code')
        else:
            # Find the existing QR code column index
            print("inside else condition")
            qr_col_idx = headers.index('qr code') + 1
        
        # Dictionary to store employees
        employees = {}
        
        # Process each employee
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Start from row 2 (1-based)
            if not row[0].value:  # Skip empty rows
                continue
                
            employee_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
            employee_id = str(employee_data.get('employee id', ''))
            
            if employee_id:
                # Generate QR code if it doesn't exist
                if not employee_data.get('qr code'):
                    qr = qrcode.QRCode(version=1, box_size=10, border=5)
                    # qr.add_data(employee_id)
                    qr_data = f"http://localhost:5000/scan/{employee_id}"  # Points to the landing page
                    qr.add_data(qr_data)
                    qr.make(fit=True)
                    img = qr.make_image(fill_color="black", back_color="white")
                    
                    # Convert to base64
                    img_io = io.BytesIO()
                    img.save(img_io, 'PNG')
                    img_io.seek(0)
                    qr_base64 = base64.b64encode(img_io.getvalue()).decode('utf-8')
                    qr_image_data = img_io.getvalue()

                    # Send email with QR code
                    send_email(
                        recipient_email=employee_data['email'],
                        employee_name=employee_data.get('name', 'Employee'),
                        qr_image_data=qr_image_data,
                        employee_id=employee_id
                    )
                    
                    # Save to Excel
                    # qr_col_idx = headers.index('qr_code') + 1  # +1 for 1-based index
                    # ws.cell(row=row_idx, column=qr_col_idx, value=qr_base64)
                    if 'qr_col_idx' not in locals():
                        qr_col_idx = headers.index('qr code') + 1  # +1 for 1-based index
                    ws.cell(row=row_idx, column=qr_col_idx, value=qr_base64)
                
                # Store in memory
                employees[employee_id] = employee_data
        
        # Save the workbook
        wb.save(app.config['EXCEL_FILE'])
        print(f"Successfully loaded {len(employees)} employees from Excel")
        return True
        
    except Exception as e:
        print(f"Error loading employees: {str(e)}")
        return False

# Load employees when the app starts
# if not load_employees():
#     print("Warning: Could not load employee data. Please check the Excel file.")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate_qr', methods=['POST'])
def generate_qr():
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        
        if not employee_id:
            return jsonify({'error': 'Employee ID is required'}), 400
            
        # Check if employee exists
        if employee_id not in employees:
            return jsonify({'error': 'Employee not found'}), 404
            
        # Generate QR code
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(employee_id)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Save to bytes
        img_io = io.BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)
        print(f"QR code generated for employee ID: {employee_id}")
        
        # return send_file(img_io, mimetype='image/png')
        response = send_file(img_io, mimetype='image/png')
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate_qr_base64', methods=['POST'])
def generate_qr_base64():
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        
        if not employee_id:
            return jsonify({'error': 'Employee ID is required'}), 400
            
        if employee_id not in employees:
            return jsonify({'error': 'Employee not found'}), 404
            
        # Generate QR code
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(employee_id)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Save to bytes
        img_io = io.BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)
        
        # Convert to base64
        img_str = base64.b64encode(img_io.getvalue()).decode('utf-8')
        
        return jsonify({
            'status': 'success',
            'employee_id': employee_id,
            'image': f'data:image/png;base64,{img_str}'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/resend_qr/<employee_id>', methods=['POST'])
def resend_qr(employee_id):
    try:
        # Load the workbook
        wb = load_workbook(filename=app.config['EXCEL_FILE'])
        ws = wb.active
        
        # Get headers
        headers = [str(cell.value).lower() if cell.value else '' for cell in next(ws.rows)]
        
        # Find the QR code column
        if 'qr code' not in headers:
            return jsonify({'error': 'QR code column not found in the Excel file'}), 400
            
        qr_col_idx = headers.index('qr code') + 1  # +1 for 1-based index
        
        # Find the employee in the Excel sheet
        employee_found = False
        employee_data = {}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header
            if not row[0].value:  # Skip empty rows
                continue
            
            print("Current employee ID:", row[0].value)
                
            current_employee_id = str(row[0].value)
            if current_employee_id == employee_id:
                employee_found = True
                # Get all employee data
                employee_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                # Get the QR code from the Excel
                qr_base64 = ws.cell(row=row_idx, column=qr_col_idx).value
                break
        
        if not employee_found or not qr_base64:
            return jsonify({'error': 'Employee not found or QR code not available'}), 404
        
        # Convert base64 back to bytes for email attachment
        qr_image_data = base64.b64decode(qr_base64)
        
        # Send email with QR code if email exists
        if 'email' in employee_data and employee_data['email']:
            send_email(
                recipient_email=employee_data['email'],
                employee_name=employee_data.get('name', 'Employee'),
                qr_image_data=qr_image_data,
                employee_id=employee_id
            )
            return jsonify({
                'status': 'success',
                'message': f'QR code resent to {employee_data["email"]}'
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'No email address found for this employee'
            }), 400
            
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to resend QR code: {str(e)}'
        }), 500


# @app.route('/employee/<employee_id>')
# def get_employee(employee_id):
#     try:
#         print("Attempting to connect to Firestore...")
#         print(f"Available collections: {[col.id for col in db.collections()]}")
        
#         employee_ref = db.collection('employees').document(employee_id)
#         print(f"Employee reference created for ID: {employee_id}")
        
#         employee = employee_ref.get()
#         print(f"Document exists: {employee.exists}")
        
#         if not employee.exists:
#             return jsonify({'error': 'Employee not found'}), 404
            
#         employee_data = employee.to_dict()
#         return jsonify(employee_data)
        
#     except Exception as e:
#         print(f"Error in get_employee: {str(e)}")
#         return jsonify({'error': str(e)}), 500


@app.route('/employee/<employee_id>')
def get_employee(employee_id):
    try:
        wb = load_workbook(filename=app.config['EXCEL_FILE'], read_only=True)
        ws = wb.active
        headers = [str(cell.value).lower() if cell.value else '' for cell in next(ws.rows)]
        
        for row in ws.iter_rows(min_row=2):  # Skip header
            if str(row[0].value) == employee_id:
                employee_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                return jsonify(employee_data)
        
        return jsonify({'error': 'Employee not found'}), 404
        
    except Exception as e:
        print(f"Error in get_employee: {str(e)}")
        return jsonify({'error': str(e)}), 500

# @app.route('/api/qr/<employee_id>')
# def get_qr_code(employee_id):
#     employee_ref = db.collection('employees').document(employee_id)
#     employee = employee_ref.get()
    
#     if not employee.exists:
#         return jsonify({'error': 'Employee not found'}), 404
    
#     employee_data = employee.to_dict()
#     return jsonify({
#         'status': 'success',
#         'employee_id': employee_id,
#         'image': f'data:image/png;base64,{employee_data.get("qr_code")}'
#     })

@app.route('/api/qr/<employee_id>')
def get_qr_code(employee_id):
    try:
        wb = load_workbook(filename=app.config['EXCEL_FILE'], read_only=True)
        ws = wb.active
        headers = [str(cell.value).lower() if cell.value else '' for cell in next(ws.rows)]
        
        if 'qr_code' not in headers:
            return jsonify({'error': 'QR code not found for any employee'}), 404
            
        qr_col_idx = headers.index('qr_code')
        
        for row in ws.iter_rows(min_row=2):  # Skip header
            if str(row[0].value) == employee_id and len(row) > qr_col_idx:
                qr_base64 = row[qr_col_idx].value
                if qr_base64:
                    return jsonify({
                        'status': 'success',
                        'employee_id': employee_id,
                        'image': f'data:image/png;base64,{qr_base64}'
                    })
        
        return jsonify({'error': 'Employee or QR code not found'}), 404
        
    except Exception as e:
        print(f"Error in get_qr_code: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Employee profile once QR is scanned
@app.route('/employee/<employee_id>')
# def employee_profile(employee_id):
#     try:
#         # Load the workbook
#         wb = load_workbook(filename=app.config['EXCEL_FILE'])
#         ws = wb.active
        
#         # Get headers
#         headers = [str(cell.value).lower() if cell.value else '' for cell in next(ws.rows)]
        
#         # Find the employee
#         for row in ws.iter_rows(min_row=2):  # Skip header
#             if not row[0].value:  # Skip empty rows
#                 continue
                
#             current_employee_id = str(row[0].value)
#             if current_employee_id == employee_id:
#                 # Map Excel data to employee dictionary
#                 employee_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                
#                 # Create employee object with all necessary fields
#                 employee = {
#                     'id': employee_id,
#                     'name': employee_data.get('name', 'Employee'),
#                     'designation': employee_data.get('designation', ''),
#                     'department': employee_data.get('department', ''),
#                     'email': employee_data.get('email', ''),
#                     'phone': employee_data.get('phone', ''),
#                     'linkedin': employee_data.get('linkedin', ''),
#                     'company_linkedin': employee_data.get('company_linkedin', ''),
#                     'bio': employee_data.get('bio', ''),
#                     'image_url': employee_data.get('image_url', '')
#                 }
#                 return render_template('employee_profile.html', employee=employee)
        
#         return jsonify({'error': 'Employee not found'}), 404
        
#     except Exception as e:
#         print(f"Error in employee_profile: {str(e)}")
#         return jsonify({'error': str(e)}), 500

# API endpoint to get employee data
@app.route('/api/employee/<employee_id>')
def get_employee_data(employee_id):
    try:
        wb = load_workbook(filename=app.config['EXCEL_FILE'])
        ws = wb.active
        headers = [str(cell.value).lower() if cell.value else '' for cell in next(ws.rows)]
        
        for row in ws.iter_rows(min_row=2):
            if not row[0].value:
                continue
                
            current_employee_id = str(row[0].value)
            if current_employee_id == employee_id:
                employee_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                return jsonify(employee_data)
        
        return jsonify({'error': 'Employee not found'}), 404
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Landing page for QR code scans
@app.route('/scan/<employee_id>')
def scan_landing(employee_id):
    return render_template('scanning.html')

# Employee profile page
@app.route('/employee/<employee_id>/profile')
def employee_profile(employee_id):
    try:
        response = get_employee_data(employee_id)
        if response.status_code != 200:
            return "Employee not found", 404
        employee = response.get_json()
        employee['id'] = employee_id
        return render_template('employee_profile.html', employee=employee)
    except Exception as e:
        return str(e), 500

@app.route('/health')
def health_check():
    return jsonify({
        'status': 'ok',
        'firebase_initialized': db is not None
    })

def _create_business_card_image(employee_id):
    """Fetches employee data and creates the business card image."""
    try:
        response = get_employee_data(employee_id)
        if response.status_code != 200:
            app.logger.warning(f"Employee with id {employee_id} not found for business card generation.")
            return None, None
        
        employee_data = response.get_json()

        # --- PNG Business Card Generation ---
        card_width, card_height = 400, 250
        padding = 20
        bg_color = (255, 255, 255)
        text_color = (0, 0, 0)
        img = Image.new('RGB', (card_width, card_height), color=bg_color)
        draw = ImageDraw.Draw(img)

        try:
            # Try common sans-serif fonts
            title_font = ImageFont.truetype("Helvetica.ttf", 20)
            details_font = ImageFont.truetype("Helvetica.ttf", 14)
        except IOError:
            try:
                title_font = ImageFont.truetype("Arial.ttf", 20)
                details_font = ImageFont.truetype("Arial.ttf", 14)
            except IOError:
                # Fallback to default font
                title_font = ImageFont.load_default()
                details_font = ImageFont.load_default()

        # Employee Details
        y_position = padding
        if employee_data.get('name'):
            draw.text((padding, y_position), employee_data.get('name', ''), font=title_font, fill=text_color)
            y_position += 30

        draw.line((padding, y_position, card_width - 140, y_position), fill='gray', width=1)
        y_position += 10

        fields_to_display = {
            'email': 'Email',
            'phone': 'Phone',
            'department': 'Department',
            'company': 'Company'
        }

        for key, label in fields_to_display.items():
            if employee_data.get(key):
                text = f"{label}: {employee_data.get(key)}"
                draw.text((padding, y_position), text, font=details_font, fill=text_color)
                y_position += 20

        # QR Code
        profile_url = url_for('employee_profile', employee_id=employee_id, _external=True)
        qr_img = qrcode.make(profile_url)
        qr_img = qr_img.resize((100, 100))
        img.paste(qr_img, (card_width - 120, card_height - 120))

        # Save image to a byte buffer
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)

        filename = f"{employee_data.get('name', 'employee').replace(' ', '_')}_card.png"
        return buf, filename

    except Exception as e:
        app.logger.error(f"Error creating business card image for {employee_id}: {e}", exc_info=True)
        return None, None

@app.route('/employee/<employee_id>/business_card')
def generate_business_card(employee_id):
    app.logger.info(f"Generating business card for employee_id: {employee_id}")
    try:
        # Create the business card image
        buf, filename = _create_business_card_image(employee_id)

        if buf and filename:
            return send_file(buf, mimetype='image/png', as_attachment=True, download_name=filename)
        else:
            return "Employee not found or error generating card", 404

    except Exception as e:
        app.logger.error(f"Error in generate_business_card route for {employee_id}: {e}", exc_info=True)
        return str(e), 500

if __name__ == '__main__':
    # Create uploads directory if it doesn't exist
    os.makedirs('static/qr_codes', exist_ok=True)
    app.run(debug=True)